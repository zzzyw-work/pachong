#!/usr/bin/env python
# -*- coding: utf-8 -*-

#导入模块 openpyxl  
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

#读取excel文件
#括号中的字符串为你要比较的两个excel的路径，注意用“/”
wb_a = openpyxl.load_workbook('e:/codefile/gitfile/pachong/pachong/before.xlsx')
wb_b = openpyxl.load_workbook('e:/codefile/gitfile/pachong/pachong/now.xlsx')
#定义一个方法来获取表格中某一列的内容，返回一个列表
def getIP(wb,index):
    sheet = wb.get_active_sheet()
    ip = []
    for cellobj in sheet[index]:
        ip.append(cellobj.value)
    return ip
#获得ip列表
# def showdiff(index):
#     ip_a = getIP(wb_a,index)
#     ip_b = getIP(wb_b,index)
#     #将两个列表转换成集合
#     aa = set(ip_a)
#     bb = set(ip_b)
#     #找出两个列表的不同行，并转换成列表
#     difference = list(aa ^ bb)

#     #return difference
#     #打印出列表中的元素
#     #到这一步，两个表格中不同的数据已经被找出来了
#     for i in difference:
#         print (i)

#将不同行高亮显示
def gaoliangdiff(index):

    ip_a = getIP(wb_a,index)
    ip_b = getIP(wb_b,index)
    #将两个列表转换成集合
    aa = set(ip_a)
    bb = set(ip_b)
    #找出两个列表的不同行，并转换成列表
    difference = list(aa ^ bb)
    #return difference
    #打印出列表中的元素
    #到这一步，两个表格中不同的数据已经被找出来了
    if difference:
        print("第%s列有差异"%index)
        for i in difference:
            print ("不同的单元内容是:%s"%i)
    else:
        print ("第%s列无差异"%index)



    

#开始高显不同的行
    print ("开始第一张表" + "----" *10)
    a = wb_a.get_active_sheet()[index]
    for cellobj in a:
        if cellobj.value in difference:
            print (cellobj.value)
            cellobj.font = Font(color=colors.BLACK, italic=True ,bold = True)
            cellobj.fill = PatternFill("solid", fgColor="DDDDDD")
    print ("开始第二张表" + "----" *10)
    b = wb_b.get_active_sheet()[index]
    for cellobj in b:
        if cellobj.value in difference:
            print (cellobj.value)
            print("\n")
            cellobj.font = Font(color=colors.BLACK, italic=True ,bold = True)
            cellobj.fill = PatternFill("solid", fgColor="DDDDDD")

#保存
    # wb_a.save('e:/codefile/gitfile/pachong/pachong/a.xlsx')
    # wb_b.save('e:/codefile/gitfile/pachong/pachong/b.xlsx')
    wb_a.save('a.xlsx')
    wb_b.save('b.xlsx')

if __name__ == '__main__':

    listqw = ['A','B','C']
    for i in listqw:
        gaoliangdiff(i)


